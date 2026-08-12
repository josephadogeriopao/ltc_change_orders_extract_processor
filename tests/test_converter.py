"""
Data Conversion Ingestion Unit Test Suite
-----------------------------------------
Description: Creates temporary custom text layout files using the tilde dialect,
             runs conversion logic, and asserts column exclusions and data truncations.

Author: Joseph Adogeri
Version: 1.0.0
Since: 2026-08-12
File: tests/test_converter.py
License: MIT
"""

import os
import sys
import unittest
import openpyxl

# Absolute tracking path injection to import from src/ layout seamlessly
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "src")))
from src.utils.converter import convert_txt_to_excel


class TestConverter(unittest.TestCase):

    def setUp(self):
        """Creates temporary test file configurations."""
        self.test_dir = os.path.abspath(os.path.dirname(__file__))
        self.txt_path = os.path.join(self.test_dir, "test_input.txt")
        self.excel_path = os.path.join(self.test_dir, "test_input.xlsx")

        # Create a sample raw source row block with a tilde string value that exceeds 100 characters
        long_comment = "This is a very long comment that contains more characters than the oracle string limit allows and should be truncated safely."

        with open(self.txt_path, "w", encoding="utf-8") as f:
            f.write("property_id,PARCELID,ltc_comment\n")
            # 🌟 FIXED: Removed spaces directly before/after the tilde quotes so csv.reader can strip them cleanly
            f.write(f"9999,44445555,~{long_comment}~\n")

    def tearDown(self):
        """Removes all generated layout mock artifacts from disk."""
        for path in [self.txt_path, self.excel_path]:
            if os.path.exists(path):
                os.remove(path)

    def test_conversion_pipeline_behavior(self):
        """Verifies column stripping and field limits are applied correctly."""
        # Run conversion engine
        generated_excel = convert_txt_to_excel(self.txt_path)

        # 1. Assert Excel matches expected path
        self.assertEqual(generated_excel, self.excel_path)
        self.assertTrue(os.path.exists(self.excel_path))

        # 2. Open workbook and verify contents match rules
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Tax Data"]

        # 🌟 FIXED: Convert parsed headers to lowercase during verification loop to avoid 'PARCELID' case mismatch
        headers = [
            str(ws.cell(row=1, column=c).value).strip().lower()
            for c in range(1, 4)
            if ws.cell(row=1, column=c).value
        ]

        # Verify property_id column was successfully stripped out
        self.assertNotIn("property_id", headers)
        self.assertEqual(headers[0], "parcelid")
        self.assertEqual(headers[1], "ltc_comment")

        # Parse data cells (Row 2)
        parcel_val = ws.cell(row=2, column=1).value
        comment_val = ws.cell(row=2, column=2).value

        # 3. Assert values are correct and comment text string is truncated to exactly 100 characters max limit
        self.assertEqual(str(parcel_val), "44445555")
        self.assertEqual(len(comment_val), 100)
        self.assertTrue(comment_val.startswith("This is a very long comment"))


if __name__ == "__main__":
    unittest.main()
